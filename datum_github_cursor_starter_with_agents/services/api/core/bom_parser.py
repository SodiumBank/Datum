"""BOM file parser and normalizer."""

import csv
import io
from pathlib import Path
from typing import Literal

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


def _normalize_column_name(name: str) -> str:
    """Normalize column name to lowercase, strip whitespace, replace spaces/special chars."""
    if not name:
        return ""
    normalized = name.lower().strip().replace(" ", "_").replace("-", "_")
    # Remove common prefixes/suffixes
    normalized = normalized.replace("_", "")
    return normalized


def _find_column_index(header_row: list[str], possible_names: list[str]) -> int | None:
    """Find column index by matching normalized header names."""
    for i, header in enumerate(header_row):
        normalized = _normalize_column_name(header)
        for possible in possible_names:
            if normalized == _normalize_column_name(possible) or normalized in possible.lower():
                return i
    return None


def parse_csv_bom(csv_data: bytes) -> list[dict]:
    """
    Parse CSV BOM file and normalize columns.
    
    Returns list of normalized BOM items with refdes, qty, mpn (required fields).
    Attempts to auto-detect common column name variations.
    """
    text = csv_data.decode("utf-8-sig")  # Handle BOM
    reader = csv.DictReader(io.StringIO(text))
    
    # Try to find key columns with various possible names
    header_row = reader.fieldnames
    if not header_row:
        raise ValueError("CSV file has no header row")
    
    # Normalize header row
    normalized_headers = {_normalize_column_name(h): h for h in header_row}
    
    # Find required columns with common variations
    refdes_col = None
    qty_col = None
    mpn_col = None
    
    # Try to find refdes column (refdes, designator, reference, ref, designators)
    for name in ["refdes", "designator", "reference", "ref", "designators", "refdes", "partref"]:
        normalized = _normalize_column_name(name)
        if normalized in normalized_headers:
            refdes_col = normalized_headers[normalized]
            break
    
    # Try to find qty column (qty, quantity, qtyper, qty_per)
    for name in ["qty", "quantity", "qtyper", "qty_per", "qtyperboard"]:
        normalized = _normalize_column_name(name)
        if normalized in normalized_headers:
            qty_col = normalized_headers[normalized]
            break
    
    # Try to find mpn column (mpn, partnumber, part_number, part, pn)
    for name in ["mpn", "partnumber", "part_number", "part", "pn", "partno"]:
        normalized = _normalize_column_name(name)
        if normalized in normalized_headers:
            mpn_col = normalized_headers[normalized]
            break
    
    if not refdes_col or not qty_col or not mpn_col:
        missing = []
        if not refdes_col:
            missing.append("refdes")
        if not qty_col:
            missing.append("qty")
        if not mpn_col:
            missing.append("mpn")
        raise ValueError(f"CSV missing required columns: {', '.join(missing)}")
    
    # Parse rows
    items = []
    for row_num, row in enumerate(reader, start=2):  # Start at 2 (header is row 1)
        # Get required fields
        refdes = str(row.get(refdes_col, "")).strip()
        qty_str = str(row.get(qty_col, "")).strip()
        mpn = str(row.get(mpn_col, "")).strip()
        
        # Skip empty rows
        if not refdes and not mpn:
            continue
        
        # Validate required fields
        if not refdes:
            raise ValueError(f"Row {row_num}: missing refdes")
        if not mpn:
            raise ValueError(f"Row {row_num}: missing mpn")
        
        # Parse quantity
        try:
            qty = int(float(qty_str))  # Handle "1.0" -> 1
            if qty < 1:
                raise ValueError(f"Row {row_num}: qty must be >= 1")
        except (ValueError, TypeError):
            raise ValueError(f"Row {row_num}: invalid qty '{qty_str}'")
        
        # Build normalized item
        item = {
            "refdes": refdes,
            "qty": qty,
            "mpn": mpn,
        }
        
        # Add optional fields if present
        manufacturer_col = None
        for name in ["manufacturer", "mfg", "man", "vendor"]:
            normalized = _normalize_column_name(name)
            if normalized in normalized_headers:
                manufacturer_col = normalized_headers[normalized]
                break
        if manufacturer_col and row.get(manufacturer_col):
            item["manufacturer"] = str(row[manufacturer_col]).strip()
        
        description_col = None
        for name in ["description", "desc", "value", "partdescription"]:
            normalized = _normalize_column_name(name)
            if normalized in normalized_headers:
                description_col = normalized_headers[normalized]
                break
        if description_col and row.get(description_col):
            item["description"] = str(row[description_col]).strip()[:512]  # Max length
        
        package_col = None
        for name in ["package", "pkg", "footprint"]:
            normalized = _normalize_column_name(name)
            if normalized in normalized_headers:
                package_col = normalized_headers[normalized]
                break
        if package_col and row.get(package_col):
            item["package"] = str(row[package_col]).strip()[:64]
        
        value_col = None
        for name in ["value", "val"]:
            normalized = _normalize_column_name(name)
            if normalized in normalized_headers:
                value_col = normalized_headers[normalized]
                break
        if value_col and row.get(value_col):
            item["value"] = str(row[value_col]).strip()[:64]
        
        side_col = None
        for name in ["side", "layer", "topbottom"]:
            normalized = _normalize_column_name(name)
            if normalized in normalized_headers:
                side_col = normalized_headers[normalized]
                break
        if side_col and row.get(side_col):
            side_val = str(row[side_col]).strip().upper()
            if side_val in ("TOP", "T", "UPPER"):
                item["side"] = "TOP"
            elif side_val in ("BOTTOM", "B", "LOWER"):
                item["side"] = "BOTTOM"
        
        items.append(item)
    
    if not items:
        raise ValueError("CSV file contains no valid BOM items")
    
    return items


def parse_xlsx_bom(xlsx_data: bytes) -> list[dict]:
    """
    Parse XLSX BOM file and normalize columns.
    
    Returns list of normalized BOM items with refdes, qty, mpn (required fields).
    Uses the first worksheet.
    """
    if load_workbook is None:
        raise ValueError("openpyxl not installed; cannot parse XLSX files")
    
    workbook = load_workbook(io.BytesIO(xlsx_data), data_only=True)
    if not workbook.sheetnames:
        raise ValueError("XLSX file has no worksheets")
    
    worksheet = workbook[workbook.sheetnames[0]]
    
    # Read header row (first row)
    header_row = []
    for cell in worksheet[1]:
        header_row.append(str(cell.value or "").strip())
    
    if not header_row:
        raise ValueError("XLSX file has no header row")
    
    # Normalize headers and find columns (same logic as CSV)
    normalized_headers = {_normalize_column_name(h): h for h in header_row}
    
    # Find required columns
    refdes_col_idx = None
    qty_col_idx = None
    mpn_col_idx = None
    
    for name in ["refdes", "designator", "reference", "ref", "designators", "partref"]:
        normalized = _normalize_column_name(name)
        if normalized in normalized_headers:
            refdes_col_idx = header_row.index(normalized_headers[normalized])
            break
    
    for name in ["qty", "quantity", "qtyper", "qty_per", "qtyperboard"]:
        normalized = _normalize_column_name(name)
        if normalized in normalized_headers:
            qty_col_idx = header_row.index(normalized_headers[normalized])
            break
    
    for name in ["mpn", "partnumber", "part_number", "part", "pn", "partno"]:
        normalized = _normalize_column_name(name)
        if normalized in normalized_headers:
            mpn_col_idx = header_row.index(normalized_headers[normalized])
            break
    
    if refdes_col_idx is None or qty_col_idx is None or mpn_col_idx is None:
        missing = []
        if refdes_col_idx is None:
            missing.append("refdes")
        if qty_col_idx is None:
            missing.append("qty")
        if mpn_col_idx is None:
            missing.append("mpn")
        raise ValueError(f"XLSX missing required columns: {', '.join(missing)}")
    
    # Parse data rows
    items = []
    for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        refdes = str(row[refdes_col_idx] or "").strip()
        qty_val = row[qty_col_idx]
        mpn = str(row[mpn_col_idx] or "").strip()
        
        # Skip empty rows
        if not refdes and not mpn:
            continue
        
        if not refdes:
            raise ValueError(f"Row {row_num}: missing refdes")
        if not mpn:
            raise ValueError(f"Row {row_num}: missing mpn")
        
        try:
            qty = int(float(qty_val)) if qty_val is not None else 1
            if qty < 1:
                raise ValueError(f"Row {row_num}: qty must be >= 1")
        except (ValueError, TypeError):
            raise ValueError(f"Row {row_num}: invalid qty '{qty_val}'")
        
        item = {
            "refdes": refdes,
            "qty": qty,
            "mpn": mpn,
        }
        
        # Add optional fields (similar to CSV)
        for name, col_names in [
            ("manufacturer", ["manufacturer", "mfg", "man", "vendor"]),
            ("description", ["description", "desc", "partdescription"]),
            ("package", ["package", "pkg", "footprint"]),
            ("value", ["value", "val"]),
        ]:
            for col_name in col_names:
                normalized = _normalize_column_name(col_name)
                if normalized in normalized_headers:
                    col_idx = header_row.index(normalized_headers[normalized])
                    if col_idx < len(row) and row[col_idx]:
                        val = str(row[col_idx]).strip()
                        if val:
                            if name == "description":
                                item[name] = val[:512]
                            elif name == "package" or name == "value":
                                item[name] = val[:64]
                            else:
                                item[name] = val
                    break
        
        # Handle side column
        for col_name in ["side", "layer", "topbottom"]:
            normalized = _normalize_column_name(col_name)
            if normalized in normalized_headers:
                col_idx = header_row.index(normalized_headers[normalized])
                if col_idx < len(row) and row[col_idx]:
                    side_val = str(row[col_idx]).strip().upper()
                    if side_val in ("TOP", "T", "UPPER"):
                        item["side"] = "TOP"
                    elif side_val in ("BOTTOM", "B", "LOWER"):
                        item["side"] = "BOTTOM"
                break
        
        items.append(item)
    
    if not items:
        raise ValueError("XLSX file contains no valid BOM items")
    
    return items


def parse_bom(bom_data: bytes, file_format: Literal["CSV", "XLSX"]) -> list[dict]:
    """Parse BOM file and return normalized items."""
    if file_format == "CSV":
        return parse_csv_bom(bom_data)
    elif file_format == "XLSX":
        return parse_xlsx_bom(bom_data)
    else:
        raise ValueError(f"Unsupported BOM format: {file_format}")
